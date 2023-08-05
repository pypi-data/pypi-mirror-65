#!/usr/bin/env python
import urllib.request
import os
import urllib.parse
import urllib.error
from xml.dom import minidom
import json
import re
from termcolor import colored
import shutil
import pandas
import webbrowser
import matplotlib.pyplot as plt
import io
import base64
import pylab as pl
import time
import datetime
import traceback
import argparse
from multiprocessing import freeze_support, Pool, Process
import ssl
import tempfile
import platform
from colorama import init
import xml.etree.ElementTree as ETree
from openpyxl import Workbook

""" Microsoft Visual C++ required, cython required for pandas installation, """
TEMP_DIR = '/tmp' if platform.system() == 'Darwin' else tempfile.gettempdir()
# Do not change these variable
RESOURCE_TYPE = "handsets"

def send_request(url):
    """send request"""
#     print("Submitting", url)
    device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
    if "All devices" in device_list_parameters or "Available devices only" in device_list_parameters:
        response = urllib.request.urlopen(url)
    else:
        response = urllib.request.urlopen(url.replace(" ", "%20"))
#    rc = response.getcode()
#    print("rc =", rc)
    return response

def send_request_with_json_response(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    maps = json.loads(text)
    return maps

def convertxmlToXls(xml, dict_keys, filename):
        '''
                Checks if file exists, parses the file and extracts the needed data
                returns a 2 dimensional list without "header"
        '''
        root = ETree.fromstring(xml)
        headers = []
        finalHeaders = []
        if dict_keys is None: 
            for child in root:
                headers.append({x.tag for x in root.findall(child.tag+"/*")})
        else:
            headers = dict_keys
        headers = headers[0]
        mdlist = []
        for child in root:
                temp = []
                for key in sorted(headers):
                    try:
                        finalHeaders.append(key)
                        temp.append(child.find(key).text)
                    except Exception as e:
                        temp.append("-")
                mdlist.append(temp)
        '''
        Generates excel file with given data
        mdlist: 2 Dimensional list containing data
        '''
        wb = Workbook()
        ws = wb.active
        for i,row in enumerate(mdlist):
                for j,value in enumerate(row):
                        ws.cell(row=i+1, column=j+1).value = value
        ws.insert_rows(0)
        #generates header
        i = 0
        finalHeaders = list(dict.fromkeys(finalHeaders))
        for i,value in enumerate(finalHeaders):
                        ws.cell(1, column=i+1).value = value
        newfilename = os.path.abspath(filename)
        wb.save(newfilename)
        return

def send_request_with_xml_response(url):
    """send request"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    xmldoc = minidom.parseString(decoded)
    return xmldoc

def send_request_to_xlsx(url, filename):
    """send_request_to_xlsx"""
    response = send_request(url)
    decoded = response.read().decode("utf-8")
    if "=list" in url:
        filename = os.path.join(TEMP_DIR, 'output', filename)
        convertxmlToXls(decoded, None, filename)

def send_request2(url):
    """send request"""
    response = send_request(url)
    text = response.read().decode("utf-8")
    return text

def get_url(resource, resource_id, operation):
    """get url """
    cloudname = os.environ['CLOUDNAME']
    url = "https://" + cloudname + ".perfectomobile.com/services/" + resource
    if resource_id != "":
        url += "/" + resource_id
    token = os.environ['TOKEN']
    if "eyJhb" in token:
        query = urllib.parse.urlencode({"operation": operation, "securityToken": token})
    else:
        if ":" not in token:
            raise Exception("Please pass your perfecto credentials in the format user:password as your second parameter!" )
        else:
            user = token.split(":")[0]
            pwd = token.split(":")[1]
            query = urllib.parse.urlencode({"operation": operation, "user": user, "password": pwd})
    url += "?" + query
    return url

def getregex_output(response, pattern1, pattern2):
    """regex"""
    matches = re.finditer(pattern1, response, re.MULTILINE)
    for match in matches:
        match_item = str(re.findall(pattern2, match.group()))
        match_item = match_item.replace(':"', "").replace('"', "")
        match_item = match_item.replace("'", "").replace("[", "")
        match_item = match_item.replace("]", "").replace(",test", "").replace(",timer.system", "").replace('description":"','')
        return str(match_item)

def device_command(exec_id, device_id, operation):
    """Runs device command"""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + "device"
    url += "&subcommand=" + operation
    url += "&param.deviceId=" + device_id
    send_request_with_json_response(url)

def end_execution(exec_id):
    """End execution"""
    url = get_url("executions/"+ str(exec_id), "", "end")
    send_request_with_json_response(url)

def start_exec():
    """start execution"""
    url = get_url("executions", "", "start")
    response = send_request2(url)
    exec_id = getregex_output(response, r'executionId\"\:\"[\w\d@.-]+\"', ":\".*$")
    return exec_id

def get_device_list_response(resource, command, status, in_use):
    """get_device_list_response"""
    url = get_url(resource, "", command)
    url += "&status=" + status
    if in_use != "":
        url += "&inUse=" + in_use
    if  len(os.environ["DEVICE_LIST_PARAMETERS"].split(":")) >= 2:
        for item in os.environ["DEVICE_LIST_PARAMETERS"].split(";"):
            if ":" in item:
                url += "&" + item.split(":")[0] + "=" + item.split(":")[1]
    xmldoc = send_request_with_xml_response(url)
    return xmldoc

def get_xml_to_xlsx(resource, command, filename): 
    """get_xml_to_xlsx"""
    url = get_url(resource, "", command)
    send_request_to_xlsx(url, filename)
    
def get_device_ids(xmldoc):
    """get_device_ids"""
    device_ids = xmldoc.getElementsByTagName('deviceId')
    return device_ids

def get_handset_count(xmldoc):
    """get_handset_count"""
    handset_elements = xmldoc.getElementsByTagName('handset')
    return len(handset_elements)

def exec_command(exec_id, device_id, cmd, subcmd):
    """exec_commands"""
    url = get_url("executions/" + str(exec_id), "", "command")
    url += "&command=" + cmd
    url += "&subcommand=" + subcmd
    url += "&param.deviceId=" + device_id
    response = send_request2(url)
    status = getregex_output(response, r'(description\"\:\".*\",\"timer.system|returnValue\"\:\".*\",\"test)', ":\".*$")
    return str(status)

def perform_actions(deviceid_color):
    """perform_actions"""
    get_network_settings = os.environ['GET_NETWORK_SETTINGS']
    deviceid_color = str(deviceid_color)
    device_id = deviceid_color.split("||",1)[0]
    color = deviceid_color.split("||",1)[1];
    desc = deviceid_color.split("||",2)[2]
    fileName = device_id + '.txt'
    file = os.path.join(TEMP_DIR, 'results', fileName)
    try:
        status = "Results="
                #update dictionary
        url = get_url(RESOURCE_TYPE, device_id, "info")
        xmldoc = send_request_with_xml_response(url)
        modelElements = xmldoc.getElementsByTagName('model')
        model = modelElements[0].firstChild.data
        osElements = xmldoc.getElementsByTagName('os')
        osDevice = osElements[0].firstChild.data
        osVElements = xmldoc.getElementsByTagName('osVersion')
        osVersion = osVElements[0].firstChild.data
        osVersion =  osDevice + " " + osVersion
        try:
            operatorElements = xmldoc.getElementsByTagName('operator')
            operator = operatorElements[0].childNodes[0].data
            phElements = xmldoc.getElementsByTagName('phoneNumber')
            phoneNumber = phElements[0].firstChild.data
        except:
            operator = "NA"
            phoneNumber ="NA"
        if "green"  in color:
            start_execution = os.environ['START_EXECUTION']
            if "true" in start_execution.lower():
                #Get execution id
                EXEC_ID = start_exec()
                #open device:
                print("opening: " + model + ", device id: " + device_id)
                device_command(EXEC_ID, device_id, "open")
                cleanup = os.environ['CLEANUP']
                if "True" in cleanup:
                    if not "iOS" in osDevice: 
                        print("cleaning up: " + model + ", device id: " + device_id)
                        status += "clean:" + exec_command(EXEC_ID, device_id, "device", "clean")
                        status += ";"
                    else:
                        status +="clean:NA;"
                reboot = os.environ['REBOOT']
                if "True" in reboot:
                    print("rebooting: " + model+ ", device id: " + device_id)
                    status += "reboot:" + exec_command(EXEC_ID, device_id, "device", "reboot")
                    status += ";"
                if "True" in get_network_settings:
                    print("getting network status of : " + model + ", device id: " + device_id)
                    networkstatus = exec_command(EXEC_ID, device_id, "network.settings", "get").replace("{","").replace("}","")
                    status += "NW:OK"
                    status += ";"
                #Close device
                device_command(EXEC_ID, device_id, "close")
                #End execution
                end_execution(EXEC_ID)
        else:
            networkstatus = ",,"

        if "True" in get_network_settings:
                final_string =  "status=" + desc + ", deviceId='" + device_id + "', model=" + str(model) + ", version=" + str(osVersion) + ", operator="+ \
                str(operator) + ", phoneNumber=" + str(phoneNumber) + ", " + str(networkstatus) + ", " + str(status)
        else:
            final_string = "status=" + desc + ", deviceId='" + device_id + "', model=" + str(model) + ", version=" + str(osVersion) + ", operator="+ \
            str(operator) + ", phoneNumber=" + str(phoneNumber) + ",,,, " + str(status)
        final_string = re.sub(r"^'|'$", '', final_string)
        f= open(file,"w+")
        f.write(str(final_string))
        f.close() 
        return final_string
    except Exception as e:
        raise Exception("Oops!" , e )
        
        if not os.path.isfile(os.path.join(TEMP_DIR, 'results', device_id + '.txt')):
            if "True" in get_network_settings:
                final_string =  "status=ERROR" + ",deviceId='" + device_id + "',,,,,,,,"
            else:
                final_string = "status=ERROR" + ",deviceId='" + device_id + "',,,,,"
            f= open(file,"w+")
            f.write(str(final_string))
            f.close() 
        return final_string

def get_list(get_dev_list):
    """get_list"""
    # Verifies each device id based on statuses
    i = 0
    split = get_dev_list.split(";")
    command = split[0]
    status = split[1]
    in_use = split[2]
    color = split[3]
    desc = split[4]
    RESPONSE = get_device_list_response(RESOURCE_TYPE, command, status, in_use)
    DEVICE_IDS = get_device_ids(RESPONSE)
    device_list = []
    if get_handset_count(RESPONSE) > 0:
        for i in range(get_handset_count(RESPONSE)):
            device_id = DEVICE_IDS[i].firstChild.data
            device_list.append(device_id + "||" + color + "||" + desc)
            device_list = [x for x in device_list if x != 0]
        if len(device_list) > 0:
            agents = get_handset_count(RESPONSE)
            pool = Pool(processes=agents)
            try:
                print("Found " + str(len(device_list)) + " devices with status: " + desc)
                output = pool.map(perform_actions, device_list)
                pool.close()  
                pool.join()
            except Exception:
                pool.close()
                pool.terminate()
                print(traceback.format_exc())
    

def fetch_details(i, exp_number, result, exp_list):
    """ fetches details"""
    if i == exp_number:
         if "=" in result:
             exp_list = exp_list.append(result.split("=", 1)[1].replace("'","").strip())
         else:
             exp_list = exp_list.append('-')
    return exp_list

def fig_to_base64(fig):
    img = io.BytesIO()
    plt.savefig(img, format='png',
                bbox_inches='tight')
    img.seek(0)
    return base64.b64encode(img.getvalue())

def print_results(results):
    """ print_results """
    i = 0
    results.sort()
    for i in range(len(results)):
        results[i]= re.sub('Results\=$','',results[i])
        results[i]= re.sub('[,]+','',results[i])
        if results[i]:
            if "Available" in results[i]:
                print(colored(results[i], "green"))
            else:
                print(colored(results[i], "red"))
        i = i + 1  
        
def prepare_html():
    """ prepare_html """
    print(colored("\nFinal Devices list:", "magenta"))
    #copies all device status to final summary
    for r, d, f in os.walk(os.path.join(TEMP_DIR , 'results')):
        for file in f:
            if ".txt" in file:
                with open(os.path.join(r, file)) as f:
                    with open(os.path.join(r, "Final_Summary.txt"), "a") as f1:
                        for line in f:
                            f1.write(line)
                            f1.write("\n")
    file = os.path.join(TEMP_DIR, 'results', 'Final_Summary.txt')
    try:
        f= open(file,"r")
    except FileNotFoundError:
        raise Exception( 'No devices found/ Re-check your arguments')
        sys.exit(-1)
    result = f.read()
    f.close() 
    print_results(result.split("\n"))
    if "true" in os.environ["PREPARE_ACTIONS_HTML"]:
        results = result.split("\n")
        #export to CSV
        new_dict = {}
        deviceids = []
        status = []
        model = []
        osVersion = []
        operator = []
        phonenumber = []
        airplanemode = []
        wifi = []
        data = []
        action_results = []
        for result in results:
            if len(result) > 0:
                new_result = result.split(",")
                new_list = []
                i = 0
                for result in new_result:
                    fetch_details(i, 0, result, status)
                    fetch_details(i, 1, result, deviceids)
                    fetch_details(i, 2, result, model)
                    fetch_details(i, 3, result, osVersion)
                    fetch_details(i, 4, result, operator)
                    fetch_details(i, 5, result, phonenumber)
                    fetch_details(i, 6, result, airplanemode)
                    fetch_details(i, 7, result, wifi)
                    fetch_details(i, 8, result, data)
                    fetch_details(i, 9, result, action_results)
                    new_list.append(result)
                    i = i + 1
        pandas.set_option('display.max_columns', None)
        pandas.set_option('display.max_colwidth', 100)
        pandas.set_option('colheader_justify', 'center') 
        get_network_settings = os.environ['GET_NETWORK_SETTINGS']
        reboot = os.environ['REBOOT']
        cleanup = os.environ['CLEANUP']
        if "True" in get_network_settings or "True" in  reboot or "True" in cleanup:
            new_dict =  {'Status': status, 'Device Id': deviceids, 'Model': model, 'OS Version': osVersion, 'Operator': operator, 'Phone number': phonenumber, 'AirplaneMode' : airplanemode, 'Wifi': wifi, 'Data': data, 'Results' : action_results}
        else:
            new_dict =  {'Status': status, 'Device Id': deviceids, 'Model': model, 'OS Version': osVersion, 'Operator': operator, 'Phone number': phonenumber}
        df = pandas.DataFrame(new_dict)
        df = df.sort_values(by ='Model')
        df = df.sort_values(by ='Status')
        df.reset_index(drop=True, inplace=True)
        pl.figure()
        pl.suptitle("Device Models")
        df['Model'].value_counts().plot(kind='barh', stacked=True)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, 'results','model.png'))
        model = '<img src="data:image/png;base64, {}"'.format(encoded.decode('utf-8'))
        pl.figure()
        pl.suptitle("Device Status")
        df['Status'].value_counts().plot(kind='barh', stacked=True)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, 'results','status.png'))
        barh = '<img src="data:image/png;base64, {}"'.format(encoded.decode('utf-8'))
        pl.figure()
        pl.suptitle("OS Versions")
        df['OS Version'].value_counts().plot(kind='barh', stacked=True)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, 'results','version.png'))
        version = '<img src="data:image/png;base64, {}"'.format(encoded.decode('utf-8'))
        pl.figure()
        pl.suptitle("SIM Operators")
        df['Operator'].value_counts().plot(kind='barh', stacked=True)
        encoded = fig_to_base64(os.path.join(TEMP_DIR, 'results','operator.png'))
        operator = '<img src="data:image/png;base64, {}"'.format(encoded.decode('utf-8'))
        df = df.sort_values(by ='Model')
        df = df.sort_values(by ='Status')
        df.to_csv(os.path.join(TEMP_DIR , 'results','output.csv'), index=False)
        current_time = datetime.datetime.now().strftime("%c")
        device_list_parameters = os.environ["DEVICE_LIST_PARAMETERS"]
        #Futuristic:
    #     le = preprocessing.LabelEncoder()
    #     #convert the categorical columns into numeric
    #     dfs = df.copy()
    #     encoded_value = le.fit_transform(dfs['Device Id'])
    #     dfs['Device Id'] = le.fit_transform(dfs['Device Id'])
    #     dfs['Status'] = le.fit_transform(dfs['Status'])
    #     dfs['Model'] = le.fit_transform(dfs['Model'])
    #     dfs['OS Version'] = le.fit_transform(dfs['OS Version'])
    #     dfs['Operator'] = le.fit_transform(dfs['Operator'])
    #     dfs['Phone number'] = le.fit_transform(dfs['Phone number'])
    #     if  "True" in get_network_settings or  "True" in reboot or  "True" in cleanup:
    #         dfs['AirplaneMode'] = le.fit_transform(dfs['AirplaneMode'])
    #         dfs['Wifi'] = le.fit_transform(dfs['Wifi'])
    #         dfs['Data'] = le.fit_transform(dfs['Data'])
    #         dfs['Results'] = le.fit_transform(dfs['Results'])
    #     print(dfs)
    #     cols = [col for col in dfs.columns if col not in ['Status','Phone number', 'OS Version', 'Model', 'Operator']]
    #     data = dfs[cols]
    #     target = dfs['Status']
    #     print(data)
    #     print(target)
        
        cloudname = os.environ['CLOUDNAME']
        html_string = '''
        <html lang="en">
          <head>
    	  <meta name="viewport" content="width=device-width, initial-scale=1">
           <meta content="text/html; charset=iso-8859-2" http-equiv="Content-Type">
    		<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/4.7.0/css/font-awesome.min.css">
            <link rel="stylesheet" href="https://www.w3schools.com/w3css/4/w3.css">
    		     <head><title>''' + cloudname.upper() + ''' Device Status Report @ ''' + current_time + '''</title>
          <script src="https://ajax.googleapis.com/ajax/libs/jquery/3.4.1/jquery.min.js"></script>
            <script>
            $(document).ready(function(){{
              $("#myInput").on("keyup", function() {{
                var value = $(this).val().toLowerCase();
                $("tbody tr").filter(function() {{
                  $(this).toggle($(this).text().toLowerCase().indexOf(value) > -1)
                }});
              }});
            }});
            </script>
    		<script type="text/javascript">
    	           $(document).ready(function(){{
                   $("#slideshow > div:gt(0)").show();
    				$("tbody tr:contains('Disconnected')").css('background-color','#fcc');
    				$("tbody tr:contains('ERROR')").css('background-color','#fcc');
    				$("tbody tr:contains('Un-available')").css('background-color','#fcc');
    				$("tbody tr:contains('Busy')").css('background-color','#fcc');
                    var table = document.getElementsByTagName("table")[0];
    				var rowCount = table.rows.length;
    				for (var i = 0; i < rowCount; i++) {{
    					if ( i >=1){{
                        available_column_number = 0;
                        device_id_column_number = 1;
    						if (table.rows[i].cells[available_column_number].innerHTML == "Available") {{
                                for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.backgroundColor = '#e6fff0';
                                        if(j=table.rows[0].cells.length){{
                                                if (table.rows[i].cells[(table.rows[0].cells.length - 1)].innerHTML.indexOf("failed") > -1) {{
                                                        table.rows[i].cells[j].style.color = '#660001';
                                                        table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
                                                }}
    							}}
                                 }}
    							var txt = table.rows[i].cells[device_id_column_number].innerHTML;
    							var url = 'https://''' + cloudname.upper() + '''.perfectomobile.com/nexperience/main.jsp?applicationName=Interactive&id=' + txt;
    							var row = $('<tr></tr>')
    							var link = document.createElement("a");
    							link.href = url;
    							link.innerHTML = txt;
    							link.target = "_blank";
    							table.rows[i].cells[device_id_column_number].innerHTML = "";
    							table.rows[i].cells[device_id_column_number].appendChild(link);
    						}}else{{
    							for(j = 0; j < table.rows[0].cells.length; j++) {{
    								table.rows[i].cells[j].style.color = '#660001';
                                         table.rows[i].cells[j].style.backgroundColor = '#FFC2B5';
    							}}
    						}}
    					}}
    				}}
                 }});
                 function myFunction() {{
                  var x = document.getElementById("myTopnav");
                  if (x.className === "topnav") {{
                    x.className += " responsive";
                  }} else {{
                    x.className = "topnav";
                  }}
                }}
    		</script>
            
    		<meta name="viewport" content="width=device-width, initial-scale=1">
            </head>
             <style>
    
            html {{
              height:100%;
            }}
    
            .mystyle {{
                font-size: 12pt;
                font-family: "Trebuchet MS", Helvetica, sans-serif;
                border-collapse: collapse;
                border: 2px solid black;
                margin-right: 2%;
                margin-left: 2%;
                box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            }}
            
            .mystyle body {{
              font-family: "Trebuchet MS", Helvetica, sans-serif;
                table-layout: auto;
                width: 100%;
                margin:0;
                position:relative;
            }}
    
            .mySlides{{
              transition:transform 0.25s ease;
            }}
           
            .mySlides:hover {{
                -webkit-transform:scale(2);
                transform:scale(2);
            }}
    
            #myInput {{
              background-image: url('http://www.free-icons-download.net/images/mobile-search-icon-94430.png');
              background-position: 2px 4px;
              background-repeat: no-repeat;
              background-size: 25px 30px;
              width: 40%;
              height:auto;
              font-weight: bold;
              font-size: 12px;
              padding: 11px 20px 12px 40px;
              box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
              transition:transform 0.25s ease;
            }}
           
            #myInput:hover {{
                -webkit-transform:scale(1.01);
                transform:scale(1.01);
            }}
    
            body {{
              background-color: #ffffff;
              background-image: linear-gradient(to right,  #09f, #bfee90, #fff, #fffdd0, #fff, #bfee90, #09f);
              height: 100%;
              top: 70%;
              background-repeat:  repeat;
              background-position: right;
              background-size:  contain;
              background-attachment: initial;
              opacity:.93;
            }}
    
            .bg {{
                background-image: linear-gradient(to bottom left, #666699 40%,  #09f 50%, #add8e6 6%, #09f 5% ) ;
              bottom:0;
              left:-50%;
              opacity:.4;
              position:absolute;
              right:-50%;
              top:0%;
              z-index:-1;
              bottom: 52%;
            }}
    
            h1 {{
              font-family:monospace;
            }}
    
            @keyframes slide {{
              0% {{
                transform:translateX(-25%);
              }}
              100% {{
                transform:translateX(25%);
              }}
            }}
    
            .mystyle table {{
                table-layout: auto;
                width: 100%;
                height: 100%;
                position:relative;
                border-collapse: collapse;
                transition:transform 0.25s ease;
            }}
           
            tr:hover {{
                -webkit-transform:scale(1.01);
                transform:scale(1.01);
            }}
    
            tr:hover {{background-color:grey;}}
    
            .mystyle td {{
                font-size: 13px;
                position:relative;
                padding: 5px;
                width:10%;
                color: black;
              border-left: 1px solid #333;
              border-right: 1px solid #333;
              background: #fffffa;
              text-align: justify;
            }}
    
            table.mystyle thead {{
              background: #333333;
              font-size: 13px;
              position:relative;
              border-bottom: 1px solid #DBDB40;
              border-left: 1px solid #D8DB40;
              border-right: 1px solid #D8DB40;
              border-top: 1px solid black;
            }}
    
            table.mystyle thead th {{
              line-height: 140%;
              font-size: 17px;
              color: white;
              text-align: center;
              transition:transform 0.25s ease;
            }}
           
            table.mystyle thead th:hover {{
                -webkit-transform:scale(1.01);
                transform:scale(1.01);
            }}
    
            table.mystyle thead th:first-child {{
              border-left: none;
              width:1%;
            }}
    
            .topnav {{
              overflow: hidden;
              background-color: #333;
              opacity: 0.7;
              background-image: linear-gradient(to right,  #666699, #013220, #333333 , #333333);
            }}
    
            .topnav a {{
              float: right;
              display: block;
              color: #333333;
              text-align: center;
              padding: 12px 15px;
              text-decoration: none;
              font-size: 12px;
              position: relative;
              border-left: 1px solid #6c3;
              border-right: 1px solid #6c3;
              transition:transform 0.25s ease;
            }}
           
            .topnav a:hover {{
                -webkit-transform:scale(1.15);
                transform:scale(1.15);
            }}
    
            .topnav a.active {{
              background-color: #333333;
              color: #b8ff7a;
              font-weight: lighter;
            }}
    
            .topnav .icon {{
              display: none;
            }}
    
            @media screen and (max-width: 600px) {{
              .topnav a:not(:first-child) {{display: none;}}
              .topnav a.icon {{
                color: #DBDB40;
                float: right;
                display: block;
              }}
            }}
    
            @media screen and (max-width: 600px) {{
              .topnav.responsive {{position: relative;}}
              .topnav.responsive .icon {{
                position: absolute;
                right: 0;
                top: 0;
              }}
              .topnav.responsive a {{
                float: none;
                display: block;
                text-align: left;
              }}
            }}
    
            footer {{
              display: block;
              font-size: 12px;
            }}
    
            * {{
              box-sizing: border-box;
            }}
    
            img {{
              vertical-align: middle;
            }}
    
            /* Position the image container */
            .container {{
              position: relative;
            }}
    
            /* Hide the images by default */
            .mySlides {{
              width: 60%;
            }}
    
            #slideshow {{
              margin:1% auto;
              position: relative;
              margin-top:5%;
              width: 60%;
              height: 82%;
              box-shadow: 0 0 80px rgba(2, 112, 0, 0.4);
            }}
    
            #ps{{
              height: 10%;
              margin-top: 0%;
              margin-bottom: 90%;
              background-position: center;
              background-repeat: no-repeat;
              background-blend-mode: saturation;
            }}
    
            #slideshow > div {{
              position: relative;
              margin-top: 10%;
              top: 10%;
              left: 1%;
              right: 1%;
              bottom: 10%;
              width: 95%;
            }}
       		#download {{
			  background-color: #333333;
			  border: none;
			  color: white;
              font-size: 12px;
              padding: 13px 20px 15px 20px;
			  cursor: pointer;
			}}

			#download:hover {{
			  background-color: RoyalBlue;
			}}
            </style>
          <div class="bg"></div>
        	<div>
          <body bgcolor="#FFFFED">
    	  	<div class="topnav" id="myTopnav">
    		  <a href="result.html" class="active">Home</a>
    		  <a href="https://''' + cloudname.upper() + '''.perfectomobile.com" target="_blank" class="active">''' + cloudname.upper() + ''' Cloud</a>
              <a href="https://developers.perfectomobile.com" target="_blank" class="active">Docs</a>
              <a href="https://www.perfecto.io/services/professional-services-implementation" target="_blank" class="active">Professional Services</a>
    		  <a href="https://support.perfecto.io/" target="_blank" class="active">Perfecto Support</a>
    		  <a href="javascript:void(0);" aria-label="first link" class="icon" onclick="myFunction()">
    			<i class="fa fa-bars"></i>
    		  </a>
    		</div>
           
            <div style="text-align: center">
            <h1> <font color=#333 ><b>''' + cloudname.upper() + ''' </h1><a href="https://''' + cloudname.upper() + '''.perfectomobile.com" target="_blank" class="site-logo">
            <img src="https://www.perfecto.io/sites/perfecto.io/themes/custom/perfecto/logo.svg" alt="Perfecto support"></a>
            <h2>Cloud's Device Status Report @ ''' + current_time + '''</font></h2></b>
    		 <input id="myInput" aria-label="search" type="text" placeholder="Search..">
             <a id ="download" href="./get_devices_list.xlsx" class="btn"><i class="fa fa-download"></i> Full Devices List</a>
             <br></p>
             <div style="overflow-x:auto;">
             {table}
             <p align="center" style="font-size:12px;font-family: "Trebuchet MS", Helvetica, sans-serif;" >Device query parameters: ''' + device_list_parameters + ''' </p> <br>
            <div class="container" align="center" id="slideshow" >
              <div class="mySlides">
                ''' + barh + ''' alt="Device Status" style="width:30%;">
              ''' + model + ''' alt="Model" style="width:30%;">
              </div>
              <div class="mySlides">
              ''' + version + ''' alt="Version" style="width:30%;">
              ''' + operator + ''' alt="Operator" style="width:30%;">
              </div>       
              </div>
            </div>
              <footer>
              <p>Best viewed in Chrome/Safari.</p>
              </footer>
          </body>
          </div>
        </html>
        '''
        
        # OUTPUT AN HTML FILE
        with open(os.path.join(TEMP_DIR,'output','result.html'), 'w') as f:
            f.write(html_string.format(table=df.to_html(classes='mystyle', index=False)))
        time.sleep(3)
        webbrowser.open('file://' + os.path.join(TEMP_DIR,'output','result.html'), new=0) 
        plt.close('all')
        print('Results: file://' + os.path.join(TEMP_DIR,'output','result.html'))
    

def create_dir(directory, delete):
    """
    create Dir
    """
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
        else:
            if delete:
                shutil.rmtree(directory)
                os.makedirs(directory)
    except Exception as e:
        print(colored(e, "red"))
        sys.exit(-1)

    
def main():
    """
    Runs the perfecto actions and reports
    """
    try:
        start_time = time.time()
        freeze_support() 
        init()
    #     """fix Python SSL CERTIFICATE_VERIFY_FAILED"""
        if (not os.environ.get('PYTHONHTTPSVERIFY', '') and getattr(ssl, '_create_unverified_context', None)):
            ssl._create_default_https_context = ssl._create_unverified_context
        parser = argparse.ArgumentParser(description="Perfecto Actions Reporter")
        parser.add_argument(
            "-c",
            "--cloud_name",
            metavar="cloud_name",
            help="Perfecto cloud name. (E.g. demo)",
        )
        parser.add_argument(
            "-s",
            "--security_token",
            metavar="security_token",
            type=str,
            help="Perfecto Security Token/ Pass your Perfecto's username and password in user:password format",
        )
        parser.add_argument(
            "-d",
            "--device_list_parameters",
            metavar="device_list_parameters",
            type=str,
            help="Perfecto get device list API parameters to limit device list. Support all API capabilities which selects devices based on reg ex/strings,  Leave it empty to select all devices",
            nargs="?"
        )
        parser.add_argument(
            "-t",
            "--device_status",
            type=str,
            metavar="Different types of Device Connection status",
            help="Different types of Device Connection status. Values: all. This will showcase all the device status like Available, Disconnected, un-available & Busy. Note: Only Available devices will be shown by default",
            nargs="?"
        )
        parser.add_argument(
            "-a",
            "--actions",
            metavar="actions",
            type=str,
            help="Perfecto actions seperated by semi-colon. E.g. reboot:true;cleanup:true;get_network_settings:true",
            nargs="?"
        )
        parser.add_argument(
            "-r",
            "--refresh",
            type=str,
            metavar="refresh",
            help="Refreshes the page with latest device status as per provided interval in seconds",
            nargs="?"
        )
        parser.add_argument(
            "-o",
            "--output",
            type=str,
            metavar="output in html",
            help="output in html. Values: true/false. Default is true",
            nargs="?"
        )
        args = vars(parser.parse_args())
        if not args["cloud_name"]:
            parser.print_help()
            parser.error("cloud_name parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoactions -c demo")
            exit
        if not args["security_token"]:
            parser.print_help()
            parser.error("security_token parameter is empty. Pass the argument -c followed by cloud_name, eg. perfectoactions -c demo -s <<TOKEN>> || perfectoactions -c demo -s <<user>>:<<password>>")
            exit
        os.environ['CLOUDNAME'] = args["cloud_name"]
        os.environ['TOKEN'] = args["security_token"]
        if args["device_list_parameters"]:
            device_list_parameters = args["device_list_parameters"]
        else:
            device_list_parameters = "All devices"
        os.environ['DEVICE_LIST_PARAMETERS'] = device_list_parameters
        os.environ['GET_NETWORK_SETTINGS'] = "False"
        reboot = "False"
        cleanup = "False"
        start_execution = "False"
        if args["actions"]:
            if "get_network_settings:true" in args["actions"]:
                os.environ['GET_NETWORK_SETTINGS'] = "True"
            if "reboot:true" in args["actions"]:
                reboot = "True"
            if "cleanup:true" in args["actions"]:
                cleanup = "True"
        os.environ["CLEANUP"] = cleanup 
        os.environ["REBOOT"] = reboot
        if "True" in os.environ['GET_NETWORK_SETTINGS'] or "True" in reboot or "True" in cleanup:
            start_execution = "True"
        os.environ["START_EXECUTION"] = start_execution
        os.environ["PREPARE_ACTIONS_HTML"] = "true"
        if args["output"]:
            if "false" in str(args["output"]).lower():
                os.environ["PREPARE_ACTIONS_HTML"] = "false" 
        os.environ["perfecto_actions_refresh"] = "false"
        if args["refresh"]:
            if int(args["refresh"]) >= 0:
                os.environ["perfecto_actions_refresh"] = args["refresh"]
        #create results path and files
        create_dir(os.path.join(TEMP_DIR , 'results'), True)
        create_dir(os.path.join(TEMP_DIR , 'output'), True)
        get_xml_to_xlsx(RESOURCE_TYPE, "list", 'get_devices_list.xlsx')
        if args["device_status"]:
            get_dev_list = ["list;connected;true;red;Busy", "list;disconnected;;red;Disconnected", \
                        "list;unavailable;;red;Un-available", "list;connected;false;green;Available"]
            try:
               procs = []
               for li in get_dev_list:
                   proc = Process(target=get_list, args=(str(li),))
                   procs.append(proc)
                   proc.start()
               for proc in procs:
                   proc.join()
               for proc in procs:
                   proc.terminate()
            except Exception:
               proc.terminate()
               print(traceback.format_exc())
               sys.exit(-1)
        else:
            if not args["device_list_parameters"]:
                os.environ['DEVICE_LIST_PARAMETERS'] = "Available Devices only"
            get_list("list;connected;false;green;Available")
            
        prepare_html()
        print("--- Completed in : %s seconds ---" % (time.time() - start_time))
        #Keeps refreshing page with expected arguments with a sleep of provided seconds   
        while "false" not in os.environ["perfecto_actions_refresh"]:
            print(str(int(os.environ["perfecto_actions_refresh"])))
            time.sleep(int(os.environ["perfecto_actions_refresh"]))
            main()
    except Exception as e:
        raise Exception("Oops!" , e )
        sys.exit(-1)

if __name__ == '__main__':
    main()
