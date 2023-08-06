from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from tools.logger import Logger
import xlrd
import xlsxwriter

lg = Logger()


class Sql(object):

    def __init__(self, user, password, port, dbname, char, base):
        self.engine = create_engine(
            'mysql://{}:{}@localhost:{}/{}?charset={}'.format(user, password, port, dbname, char),
            echo=False)
        self.__Base = base
        self.__Base.metadata.create_all(self.engine)
        self.__Session = sessionmaker(bind=self.engine)
        self.__session = self.__Session()

    def write(self, msg):
        self.__session.add(msg)
        self.__session.commit()

    @property
    def session(self):
        return self.__session


class ExcelWriter(object):
    def __init__(self, path):
        self.work_book = xlsxwriter.Workbook(path)  # 创建一个excel文件
        self.sheet_list = []

    def create_sheet(self, sheet_name='sheet'):
        self.sheet_list.append(self.work_book.add_worksheet(sheet_name))  # 在文件中创建一个名为TEST的sheet,不加名字默认为sheet1

    def write_line(self, line, row_num=0, col_num=0, sheet_index=0):
        self.sheet_list[sheet_index].write_row(row_num, col_num, line)

    def write_lines(self, lines, sheet_index=0):
        if not self.sheet_list:
            self.create_sheet()
        for i, d in enumerate(lines):
            self.sheet_list[sheet_index].write_row(i, 0, d)

    def write_cols(self, *args, sheet_index=0):
        col_lenght = len(args)
        tag = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        tag_list = list(tag[:col_lenght])
        for i, d in enumerate(args):
            self.sheet_list[sheet_index].write_column('{}{}'.format(tag_list[i], 2), d)

    def close_book(self):
        self.work_book.close()


class ExcelReader(object):
    def __init__(self, path=None):
        self.sheet_list = []
        self.name = path.split('/')[-1].split('.')[0]
        self.data = []
        if path:
            self.read_excel(path)

    def read_excel(self, path):
        wb = xlrd.open_workbook(filename=path)  # 打开文件
        tb_list = wb.sheet_names()  # 获取所有表格名字
        wb_num = len(tb_list)

        # get sheet object list
        for i in range(wb_num):
            sheet = wb.sheet_by_index(i)  # 通过索引获取表格
            self.sheet_list.append(sheet)

        # get data
        for i in self.sheet_list:
            table_data = []
            tabel_length = len(i.col_values(0))
            for j in range(tabel_length):
                table_data.append(i.row_values(j))  # 获取行内容
            self.data.append(table_data)

        return self.data

    def to_csv(self, path=None, sheet_index=0):
        if not path:
            path = self.name + '.csv'
        csv = CsvWriter(path)
        csv.write_lines(self.data[sheet_index])


class CsvWriter(object):

    def __init__(self, path=None):
        self.save_path = path
        self.name = self.save_path.split('/')[-1].split('.')[0]

    def set_save_path(self, path):
        self.save_path = save_path

    def write_line(self, line):
        # ['1','2','3']
        with open(self.save_path, 'a+') as f:
            line = [str(i) for i in line]
            f.write(','.join(line) + '\n')

    @lg.log()
    def write_lines(self, lines):
        # [['1','2','3'], [...], ...]
        with open(self.save_path, 'a+') as f:
            for i in lines:
                line = ','.join([str(j) for j in i])
                f.write(line + '\n')
                f.flush()
            lg.info('File \'{}\' Write Done!'.format(self.save_path))

    @lg.log()
    def write_cols(self, *args):
        # [['headline','1','2','3'], [...], ...]
        with open(self.save_path, 'a+') as f:
            for i in zip(*args):
                line = ','.join(i) + '\n'
                f.write(line)
                f.flush()
            lg.info('Write Done!')


class CsvReader(object):
    def __init__(self, path=None):
        self.data = []
        self.name = path.split('/')[-1].split('.')[0]
        if path:
            self.read_csv(path)

    def read_csv(self, path):
        with open(path, 'r') as f:
            for i in f:
                self.data.append(i.strip())

        return self.data

    @lg.log()
    def to_execl(self, path=None):
        if not path:
            path = self.name + '.xlsx'
        ew = ExcelWriter(path)
        lines = [i.split(',') for i in self.data]
        ew.write_lines(lines)
        ew.close_book()
        lg.info('File \'{}\' Write Done!'.format(path))

