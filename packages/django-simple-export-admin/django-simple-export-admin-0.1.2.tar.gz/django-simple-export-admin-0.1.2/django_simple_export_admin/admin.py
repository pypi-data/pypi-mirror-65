import urllib
from django.contrib import admin
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.urls import path
from django.urls import reverse
from django.contrib import messages
from django.utils.translation import ugettext_lazy as _
from django.utils import timezone
from django.utils.http import urlquote
from django_changelist_toolbar_admin.admin import DjangoChangelistToolbarAdmin
from django_changelist_toolbar_admin.admin import Button
import openpyxl

DEFAULT_FILENAME_TEMPLATE = "{model_name}-{year:04d}{month:02d}{day:02d}-{hour:02d}{minute:02d}{second:02d}.xlsx"


class Aggregate(object):
    pass

class Sum(Aggregate):
    def __init__(self):
        self.total = 0
    
    def push(self, value):
        self.total += value
    
    def final(self):
        return self.total

class Average(Aggregate):
    def __init__(self, empty="-"):
        self.empty = empty
        self.total = 0
        self.count = 0

    def push(self, value):
        self.total += value
        self.count += 1
    
    def final(self):
        if not self.count:
            return empty
        else:
            return self.total / self.count

class Count(Aggregate):
    def __init__(self, value):
        self.value = value
        self.count = 0
    
    def push(self, value):
        if self.value == value:
            self.count += 1
    
    def final(self):
        return self.count

class DateRender(object):
    def __init__(self, format="%Y/%m/%d", empty_value="-"):
        self.format = format
        self.empty_value = empty_value

    def __call__(self, value):
        if not value:
            return self.empty_value
        else:
            return value.strftime(self.format)

class BooleanRender(object):
    def __init__(self, true_display="TRUE", false_display="FALSE"):
        self.true_display = true_display
        self.false_display = false_display
    
    def __call__(self, value):
        if value:
            return self.true_display
        else:
            return self.false_display

class NullBooleanRender(object):
    def __init__(self, null_display="NULL", true_display="TRUE", false_display="FALSE"):
        self.null_display = null_display
        self.true_display = true_display
        self.false_display = false_display
    
    def __call__(self, value):
        if value is None:
            return self.null_display
        if value:
            return self.true_display
        else:
            return self.false_display

class DjangoSimpleExportAdmin(DjangoChangelistToolbarAdmin, admin.ModelAdmin):

    def get_urls(self):
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        return [
            path("simple-export/<code>/", self.admin_site.admin_view(self.django_simple_export_admin_export_view), name="{0}_{1}_simple_export".format(app_label, model_name)),
        ] + super().get_urls()

    def django_simple_export_admin_export_view(self, request, code):
        django_simple_exports = self.django_simple_export_admin_get_django_simple_exports()
        settings = django_simple_exports.get(code, None)
        if not settings:
            msg = _("No export setings for: %s.") % code
            return self.django_simple_export_admin_redirect_to_changelist_with_error_message(request, msg)
        if not self.django_simple_export_admin_has_perm(request, settings.get("permissions", [])):
            msg = _("No export permission.")
            return self.django_simple_export_admin_redirect_to_changelist_with_error_message(request, msg)
        try:
            return self.django_simple_export_admin_do_export(request, settings)
        except Exception as err:
            msg = _("Something failed while do exporting: %s.") % (str(err))
            return self.django_simple_export_admin_redirect_to_changelist_with_error_message(request, msg)

    def django_simple_export_admin_redirect_to_changelist_with_error_message(self, request, msg):
        messages.add_message(request, messages.ERROR, msg)
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        url = reverse("admin:{}_{}_changelist".format(app_label, model_name))
        return HttpResponseRedirect(url)

    def django_simple_export_admin_get_django_simple_exports(self):
        if hasattr(self, "django_simple_export_admin_exports"):
            return self.django_simple_export_admin_exports
        else:
            return {}

    def django_simple_export_admin_get_filename(self, request, settings):
        filename_template = settings.get("filename", DEFAULT_FILENAME_TEMPLATE)
        now_time = timezone.localtime(timezone.now())
        info = {
            "app_label": self.model._meta.app_label,
            "model_name": self.model._meta.model_name,
            "app_verbose_name": "",
            "model_verbose_name": self.model._meta.verbose_name,
            "model_verbose_name_plural": self.model._meta.verbose_name_plural,
            "now_time": now_time,
            "year": now_time.year,
            "month": now_time.month,
            "day": now_time.day,
            "hour": now_time.hour,
            "minute": now_time.minute,
            "second": now_time.second,
            "datetime": now_time.strftime("%Y%m%d-%H%M%S"),
            "timestamp": now_time.timestamp(),
        }
        return filename_template.format(**info)

    def django_simple_export_admin_do_export(self, request, settings):
        filename = self.django_simple_export_admin_get_filename(request, settings)
        response = HttpResponse(content_type="application/ms-excel")
        response["Content-Disposition"] = """attachment; filename="{0}" """.format(urlquote(filename)).strip()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        footer_values = {}
        row_index = 0
        for col_index, field_setting in enumerate(settings["fields"]):
            worksheet.cell(row_index+1, col_index+1).value = field_setting.get("label", field_setting.get("field", ""))
        row_index += 1
        for item in self.django_simple_export_admin_get_items(request):
            for col_index, field_setting in enumerate(settings["fields"]):
                value = self.django_simple_export_admin_get_field_value(item, field_setting, row_index)
                worksheet.cell(row_index+1, col_index+1).value = value
                self.django_simple_export_admin_calc_footer_value(footer_values, col_index, field_setting, value)
            row_index += 1
        for col_index, field_setting in enumerate(settings["fields"]):
            value = footer_values[col_index]
            if isinstance(value, Aggregate):
                value = value.final()
            if value:
                worksheet.cell(row_index+1, col_index+1).value = value
        row_index += 1
        workbook.save(response)
        return response

    def django_simple_export_admin_calc_footer_value(self, footer_values, col_index, field_setting, value):
        if not col_index in footer_values:
            field_value = field_setting.get("footer-value", None)
            if field_value and callable(field_value):
                field_value = field_value()
            footer_values[col_index] = field_value
        if isinstance(footer_values[col_index], Aggregate):
            footer_values[col_index].push(value)

    def django_simple_export_admin_get_items(self, request):
        cl = self.get_changelist_instance(request)
        return cl.result_list

    def django_simple_export_admin_get_field_value(self, item, settings, row_index):
        field_name = settings["field"]
        if field_name.lower() == "__row_index":
            field_value = row_index
        else:
            if hasattr(item, field_name):
                field_value = getattr(item, field_name)
                if callable(field_value):
                    field_value = field_value()
            elif hasattr(self, field_name):
                field_value = getattr(self, field_name)(item)
            else:
                raise KeyError("field {0} not exists...".format(field_name))
            if "render" in settings:
                field_value = settings["render"](field_value)
            elif "empty_value" in settings and field_value is None:
                field_value = settings["empty_value"]
        return field_value

    def get_changelist_toolbar_buttons(self, request):
        buttons = super().get_changelist_toolbar_buttons(request)
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name
        for code, settings in self.django_simple_export_admin_get_django_simple_exports().items():
            if not self.django_simple_export_admin_has_perm(request, settings.get("permissions", [])):
                continue
            export_filtered = settings.get("export-filtered", False)
            url = reverse("admin:{0}_{1}_simple_export".format(app_label, model_name), args=[code])
            if export_filtered:
                url += "?" + request.GET.urlencode()
            title = settings.get("label", _("Export"))
            icon = settings.get("icon", None)
            klass = settings.get("class", None)
            button = Button(url, title, klass=klass, icon=icon)
            buttons.append(button)
        return buttons

    def django_simple_export_admin_has_perm(self, request, permissions):
        if request.user.is_superuser:
            return True
        for permission in permissions:
            if request.user.has_perm(permission):
                return True
        return False
