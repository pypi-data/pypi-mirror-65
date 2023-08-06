from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


class Excel(object):

    def __init__(self, excel_path, new=False):
        '''
            创建EXCEL
        :param excel_path:EXCEL目录
        :param new:是否重新创建
        '''
        self.excel_path = excel_path
        if new:
            self.wb = Workbook()
            self.wb.save(excel_path)
        self.wb = load_workbook(excel_path)
        self.MAP = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

    def set_col_width(self, excel_path, excel_sheet, col_list=[]):
        '''
            设置列宽度
        :param excel_path:excel目录
        :param excel_sheet:sheet名称
        :param col_list:要改的col列表
        :return:
        '''
        ws = self.get_sheet(excel_path, excel_sheet)
        for i in range(len(col_list)):
            ws.column_dimensions[self.MAP[i]].width = col_list[i] * 2
        if type(excel_path) == str and type(excel_sheet) == str:
            self.wb.save(excel_path)  # 保存文件

    def set_row_colour(self, excel_path, excel_sheet, row=None, colour='E1FFFF'):
        '''
            为单元格填充颜色
        :param excel_path:excel目录
        :param excel_sheet:sheet名称
        :param row:第几行
        :param colour:颜色值
        :return:
        '''
        fill = PatternFill(patternType="solid", start_color=colour)
        ws = self.get_sheet(excel_path, excel_sheet)
        if row is None:
            row = ws.max_row
        for i in range(ws.max_column):
            ws['{}{}'.format(self.MAP[i], row)].fill = fill
        if type(excel_path) == str and type(excel_sheet) == str:
            self.wb.save(excel_path)  # 保存文件

    def get_end_value(self, excel_path, excel_sheet, max_line=None):
        '''
            整行读取值
        :param excel_path:excel目录
        :param excel_sheet:sheet名称
        :param max_line:查找行
        :return:
        '''
        ws = self.get_sheet(excel_path, excel_sheet)
        list1 = []
        rows = ws.rows
        for row in rows:
            line = [col.value for col in row]
            list1.append(line)
        if max_line is None:
            max_line = ws.max_row - 1
        return list1[max_line]

    def insert_image(self, excel_path, excel_sheet, image_path, col, row=None):
        '''
            插入图片
        :param excel_path:excel目录
        :param excel_sheet:sheet名称
        :param image_path:image目录
        :param col:列
        :param row:行
        :return:
        '''
        ws = self.get_sheet(self.excel_path, excel_sheet)
        ws.column_dimensions[self.MAP[col]].width = 13  # 修改列宽
        if row is None:
            row = ws.max_row
        ws.row_dimensions[row].height = 36  # 修改行高
        img = Image(image_path)  # 获取图像
        img.width, img.height = 100, 50  # 图片设置宽高
        ws.add_image(img, '{}{}'.format(self.MAP[col], row))  # 插入图片
        if type(excel_path) == str and type(excel_sheet) == str:
            self.wb.save(excel_path)  # 保存文件

    def add_value(self, excel_path, excel_sheet, list_value):
        '''
            添加一列值
        :param excel_path:excel路径
        :param excel_sheet:sheet名称
        :param list_value:插入列表值
        :return:
        '''
        ws = self.get_sheet(self.excel_path, excel_sheet)
        ws.append(list_value)
        self.wb.save(excel_path)

    def get_sheet(self, excel_path, excel_sheet):
        '''
            获取Sheet
        :param self.wb:打开的excel
        :param excel_sheet:sheet名称
        :return:ws
        '''
        if type(excel_path) == str and type(excel_sheet) == str:
            self.wb = load_workbook(excel_path)
        else:
            self.wb = excel_path
        if len(self.wb.sheetnames) < 2 and self.wb.sheetnames[0] == "Sheet":
            ws = self.wb.active
            ws.title = excel_sheet
        else:
            if type(excel_sheet) == str:
                if not excel_sheet in self.wb.sheetnames:  # 遍历所有sheet
                    self.wb.create_sheet(title=excel_sheet)
                ws = self.wb[excel_sheet]
            else:
                ws = excel_sheet
        return ws

    def end_save(self, excel_path, excel_sheet):
        '''
            最后设置整体EXCEL格式
        :param excel_path:excel目录
        :return:
        '''
        self.wb.save(excel_path)  # 先保存之前数据
        ws = self.get_sheet(excel_path, excel_sheet)
        font = Font(size=11, bold=False, name='微软雅黑', color="000000")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # wrap_text=True 打开自动换行
        max_row = ws.max_row
        max_col = ws.max_column
        for row in ws['A1:{}{}'.format(self.MAP[max_col - 1], max_row)]:
            for cell in row:
                cell.font = font  # 修改文字颜色
                cell.border = border  # 加边框
                cell.alignment = alignment  # 修改格式
                if cell.value == "Fail" or cell.value == "fail":
                    cell.fill = PatternFill(patternType="solid", start_color="FF0000")
        for i in range(ws.max_row + 1):
            ws.row_dimensions[i].height = 36
        self.wb.save(excel_path)

    def other(self, ws):
        '''
            没用上资料备份
        :param ws: 打开的sheet
        :return:
        '''
        ws.sheet_properties.tabColor = "1072BA"  # 为sheet底部加颜色
        ws.append(['aa', 'ss'])  # 写入整column
        rows = ws.rows  # 读取整row
        columns = ws.columns  # 读取整column
        rows = ws.max_row  # 获取最大行数
        cols = ws.max_column  # 获取最大列数
        ws['A4'] = 4  # 单行写入
        print(ws['A4'].value)  # 单行读取
        ws.column_dimensions['A'].width = 13  # 修改列宽
        ws.row_dimensions[1].height = 36  # 修改行高
        # 合并单元格
        ws.merge_cells('A1:G1')
        ws.save(path)


if __name__ == "__main__":
    path = r'./excel.xlsx'
    excel = Excel(path, True)
    ws = excel.get_sheet(path, '计划')
    # excel.insert_image(path, '计划', './1.png', 'A', 3)
    # print(excel.get_send_value(path, '计划'))
    ws['B5'] = "Fail"
    excel.wb.save(path)
    excel.set_col_width(path, '计划', [16, 17, 16])
    excel.end_save(path, '计划')
