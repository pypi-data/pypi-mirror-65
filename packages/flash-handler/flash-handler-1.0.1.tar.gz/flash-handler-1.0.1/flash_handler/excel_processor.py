import os

import matplotlib
from matplotlib import font_manager
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd

from flash_handler.const import Const

#
# matplotlib的FontManager在实例化过程中可能会打日志。例如：
#   Could not open font file /System/Library/Fonts/Supplemental/NISC18030.ttf: In FT2Font: Could not set the fontsize
# 这些日志是用logging.INFO级别打印的。
# handler库不希望用户看到这些日志，所以提高日志级别。
#
matplotlib.set_loglevel('error')


def _setup_preferred_font():
    """设置合适的字体，使中文字符在图表中能正常显示。该函数优先使用沙盒中存在的中文字体：
    `Source Han Sans CN`。"""
    preferred_font = 'Source Han Sans CN'

    fm = font_manager.FontManager()

    font_path = None
    for font in fm.ttflist:
        if font.name == preferred_font:
            font_path = font.fname
            break
    else:
        return False

    font_dir = os.path.sep.join(font_path.split(os.path.sep)[:-1])
    font_files = font_manager.findSystemFonts(fontpaths=[font_dir])
    font_list = font_manager.createFontList(font_files)
    font_manager.fontManager.ttflist.extend(font_list)

    plt.rcParams['font.family'] = preferred_font
    return True


def _setup_remedial_font():
    """设置合适的字体，使中文字符在图表中能正常显示。该函数使用常见的支持中文的字体的情况。"""
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS']


def _setup_matplotlib():
    """初始化matplotlib。例如，设置支持中文字符的字体等。"""

    # 让matplotlib在绘制负号时使用ascii字符集的负号而不是unicode的负号
    plt.rcParams['axes.unicode_minus'] = False

    # 指定字体：解决plot不能显示中文问题
    if not _setup_preferred_font():
        _setup_remedial_font()


_setup_matplotlib()


def _float_or_str(str_):
    try:
        return float(str_)
    except ValueError:
        return str_


def _guess_and_convert_type(str_):
    """猜测并转换字符串至合理的类型：
    - 如果传入参数是整数字符串，就认为它是浮点数
    - 否则，如果传入参数是浮点字符串，就认为它是整数
    - 否则，保持不变。
    注：该函数假定传入参数是字符串类型。"""
    try:
        return int(str_)
    except ValueError:
        return _float_or_str(str_)


class ExcelProcessor(object):
    """用于操作单个Excel文件的基础模块，不存在与用户交互。"""

    @staticmethod
    def search(file_path, primary_key, secondary_key):
        """根据主键primary_key和次键secondary_key，搜索位于路径file_path的Excel文件，
        并将搜索结果返回。"""
        if not (primary_key or secondary_key):
            return []

        df = pd.read_excel(file_path)

        # 在primary_key为数值字符串时，primary_key_float_or_str会是浮点类型。
        # 从而在excel中的数据存储为浮点类型时，把数值相等的项目匹配出来。
        primary_key_float_or_str = _float_or_str(primary_key)
        df = df[(df.values == primary_key) |
                (df.values == primary_key_float_or_str)]

        if secondary_key:
            secondary_key_float_or_str = _float_or_str(secondary_key)
            df = df[(df.values == secondary_key) |
                    (df.values == secondary_key_float_or_str)]

        return df

    @staticmethod
    def append(file_path, values):
        """向位于路径file_path的Excel文件添加一行新的数据，数据具体取值由values决定。"""
        values = [_guess_and_convert_type(value) for value in values]

        workbook = openpyxl.load_workbook(file_path)
        current_sheet = workbook.active
        current_sheet.append(values)
        workbook.save(file_path)

    @staticmethod
    def delete(file_path, row_indexes):
        """从位于路径file_path的Excel文件中删除一些行，具体哪些行由行的索引row_indexes
        指定。"""
        workbook = openpyxl.load_workbook(file_path)
        current_sheet = workbook.active

        # 从高位开始删除免得行数变化
        row_indexes = sorted(list(row_indexes), reverse=True)
        for row_index in row_indexes:
            # excel从1开始计数；注意第一行为表头
            current_sheet.delete_rows(row_index + 2, 1)
        workbook.save(file_path)

    @staticmethod
    def update(file_path, row_indexes, column_index, new_value):
        """修改位于路径file_path的Excel文件中的一些行，将这些行的指定列的取值更改为
        new_value。具体哪些行由行的索引row_indexes指定，具体哪一列由列的索引column_index
        指定。"""
        workbook = openpyxl.load_workbook(file_path)
        current_sheet = workbook.active
        new_value = _guess_and_convert_type(new_value)
        for row_index in row_indexes:
            # excel从1开始计数；注意第一行为表头
            cell = current_sheet.cell(row=row_index+2, column=column_index+1)
            cell.value = new_value
        workbook.save(file_path)

    @staticmethod
    def read_all(file_path):
        """读取位于路径file_path的Excel文件的所有数据并返回。"""
        return pd.read_excel(file_path, dtype=object)

    @staticmethod
    def read_headers(file_path):
        """读取位于路径file_path的Excel文件的表头并返回。"""
        return list(pd.read_excel(file_path))

    @classmethod
    def draw_chart(cls, file_path, chart_type, horizontal_axis, vertical_axis,
                   method, title, output_path):
        """绘制图表。

        :param file_path: 提供数据的Excel文件的路径
        :param chart_type: 图表类型：柱状图、饼图、折线图
        :param horizontal_axis: Excel文件中用作横坐标的列名
        :param vertical_axis: Excel文件中用来计算图表纵坐标值（柱状图或折线图）或统计值
                              （饼图）的列的名字
        :param method: 计算纵坐标值（柱状图或折线图）或统计值（饼图）的方法：计数或加和
        :param title: 图表标题
        :param output_path: 图表保存的路径
        """
        df = pd.read_excel(file_path)

        if Const.METHOD_SUM_TYPE == method:
            df = df.groupby(horizontal_axis).agg(np.sum)
        else:
            assert(Const.METHOD_COUNT_TYPE == method)
            df = df.groupby(horizontal_axis).count()

        axis = df[vertical_axis].plot(kind=chart_type, title=title)

        axis.set_xlabel(horizontal_axis)
        axis.set_ylabel(vertical_axis)

        if Const.CHART_LINE_TYPE == chart_type:
            cls._post_draw_line_chart(df, axis)

        fig = axis.get_figure()
        fig.savefig(output_path)
        fig.clear()

    @staticmethod
    def _post_draw_line_chart(df, axis):
        """折线图特殊处理，用于调整折线图的刻度显示，使其与柱状图在样式上风格一致。

        :param df: 数据源，是pandas DataFrame类型。
        :param axis: matplotlib轴对象。"""

        # 折线图x轴刻度较多时，默认会选择性显示刻度，此处将所有刻度完全显示。
        plt.xticks(range(0, len(df.index)), df.index)

        # 旋转刻度，与柱状图保持一致
        axis.set_xticklabels(df.index, rotation=90)
