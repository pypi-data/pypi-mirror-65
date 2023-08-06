# -*- coding: utf-8 -*-

"""Utils for sanity check."""

from collections import defaultdict

from openpyxl import load_workbook


def munge_labels(label):
    """Munge labels."""
    remove_set = ['*', ' ', '|', '-', '"', "'", "\n", "↑", "↓"]
    split_set = ['/', ',']

    label = str(label).lower()

    for symb in remove_set:
        if symb in label:
            label = label.replace(symb, '')

    if '/' in label:
        label = tuple(set(label.split('/')))
        if len(label) == 1:
            label = label[0]

    if ',' in label:
        label = tuple(set(label.split(',')))
        if len(label) == 1:
            label = label[0]

    return label


def parse_set1(path):
    """Process dataset 1."""
    wb = load_workbook(filename=path)
    sheet_titles = []
    omics_labels = defaultdict(lambda: set())

    for sheet in wb:
        cell_value = sheet['A3'].value

        # if "Expression data (FC) of the differentially expressed" in sheet['A1'].value:
        #    sheet_title = sheet['A1'].value.split("Expression data (FC) of the differentially expressed ",1)[1]
        #    sheet_title = sheet_title.split(" of HepG2 cells after treatment with ")
        #    sheet_title[1] = sheet_title[1].replace(". Statistical significance (p value < 0.05) is indicated.", "").replace(" CsA for", "")
        #    sheet_titles.append(sheet_title)

        if cell_value and ("Significant " in cell_value or "Metabolite" == cell_value):
            if cell_value == "Metabolite":
                sheet_title = ("Metabolite", '3 µM', ' 24h or 72h')
                min_row = 3

            else:
                sheet_title = cell_value.split("Significant ", 1)[1]
                sheet_title = sheet_title.split(" CsA ")
                sheet_title.append(sheet_title[1].split(" ")[0] + ' h')
                sheet_title[1] = sheet_title[0].split(" ")[1] + ' µM'
                sheet_title[0] = sheet_title[0].split(" ")[0]
                min_row = 4

            for col in sheet.iter_cols(min_row=min_row):
                col_label = col[0].value
                sheet_omic = sheet_title[0]

                if col_label in ['MicroRNA', 'hgnc symbol', 'Metabolite']:
                    # for cell in col[1:]:
                    # if munge_labels(cell.value) == '':
                    # print(cell)
                    # print(cell.value)
                    omics_labels[sheet_omic.lower()].update(
                        munge_labels(cell.value) for cell in col[1:] if munge_labels(cell.value) != '')

            sheet_titles.append(sheet_title)

    return omics_labels


def parse_set2(path):
    """Process dataset 2."""
    wb = load_workbook(filename=path)
    omics_labels = defaultdict(lambda: set())

    for sheet in wb:
        cell_value = sheet['A1'].value

        if "metabolites" in cell_value:
            sheet_title = "metabolite"
            min_row = 2

        elif "genes" in cell_value:
            sheet_title = "genes"
            min_row = 2

        else:
            continue

        for col in sheet.iter_cols(min_row=min_row):
            col_label = col[0].value

            if col_label in ['Metabolites', 'Metabolites name', 'Genes', 'microRNA']:
                if col_label == 'microRNA':
                    col_label = 'micrornas'

                if col_label == 'Metabolites':
                    col_label = 'metabolite'

                omics_labels[col_label.lower()].update(
                    munge_labels(cell.value) for cell in col[1:] if munge_labels(cell.value) != '')

    return omics_labels


def parse_set3(path):
    """Process dataset 3."""
    wb = load_workbook(filename=path)
    omics_labels = defaultdict(lambda: set())

    for sheet in wb:
        cell_value = sheet['A1'].value

        # if "Expression data (FC) of the differentially expressed" in sheet['A1'].value:
        #    sheet_title = sheet['A1'].value.split("Expression data (FC) of the differentially expressed ",1)[1]
        #    sheet_title = sheet_title.split(" of HepG2 cells after treatment with ")
        #    sheet_title[1] = sheet_title[1].replace(". Statistical significance (p value < 0.05) is indicated.", "").replace(" CsA for", "")
        #    sheet_titles.append(sheet_title)

        if "compounds" in cell_value:
            sheet_title = "metabolite"
            min_row = 2

        elif "mRNAs" in cell_value:
            sheet_title = "genes"
            min_row = 2

        elif "miRNAs selected" in cell_value:
            sheet_title = "micrornas"
            min_row = 2

        elif "pathway" in cell_value:
            sheet_title = "pathway"
            min_row = 2

        else:
            continue

        for col in sheet.iter_cols(min_row=min_row):
            col_label = col[0].value

            if col_label in ['Name', 'Abbreviation', 'miRNA']:
                omics_labels[sheet_title].update(
                    munge_labels(cell.value) for cell in col[1:] if munge_labels(cell.value) != '')

    return omics_labels
