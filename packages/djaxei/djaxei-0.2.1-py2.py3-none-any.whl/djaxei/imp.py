import os
from tempfile import NamedTemporaryFile

import xlsxwriter
from django.contrib.admin.utils import NestedObjects
from django.db import router
from django.utils.translation import ugettext_lazy as _
from openpyxl import load_workbook


class Importer:
    def __init__(self, tmpdir=None, **kwargs):
        self.tmpdir = tmpdir

    def xls_import(self, file,  models=[], *args, **kwargs):
        models = [x.lower() for x in models]
        wb = load_workbook(file)
        worksheets = {x: wb[x.split('.')[-1]] for x in models}
        records = {}
        for model_fqn, ws in worksheets.items():
            for row in ws.iter_rows(values_only=True):
                print(model_fqn, ws.title, row)
